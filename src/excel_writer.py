import pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter

def update_excel(data, file_path="data/solar_return.xlsx", date_range=None, sheet_name="Sheet1", baseline_rate=None):
    # Use provided baseline_rate or import default from settings
    if baseline_rate is None:
        from config.settings import BASELINE_RATE
        baseline_rate = BASELINE_RATE
    if date_range is None:
        date_range = datetime.now().strftime("%Y-%m")
    
    # Format data before writing to Excel
    formatted_data = {
        "date_range": date_range,
    }
    
    for key, value in data.items():
        if "kwh" in key.lower():
            # Format kWh to 1 decimal place
            formatted_data[key] = round(value, 1)
        elif "cost" in key.lower() or "income" in key.lower() or "savings" in key.lower() or "return" in key.lower():
            # Format currency to 2 decimal places
            formatted_data[key] = round(value, 2)
        else:
            formatted_data[key] = value
    
    df = pd.DataFrame([formatted_data])

    try:
        existing = pd.read_excel(file_path)
        # Remove duplicate columns (actual_cost, no_solar_cost, net_returns) from existing data
        # Keep only the first occurrence of each
        cols_to_check = ['actual_cost', 'no_solar_cost', 'net_returns']
        for col in cols_to_check:
            # Find all columns with this name (including pandas' auto-renamed ones like 'actual_cost.1')
            cols_to_drop = [c for c in existing.columns if c == col or c.startswith(col + '.')]
            if len(cols_to_drop) > 1:
                # Keep the first, drop the rest
                existing = existing.drop(columns=cols_to_drop[1:])
        
        df = pd.concat([existing, df], ignore_index=True)
    except FileNotFoundError:
        pass

    # Write to Excel with formatting
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        
        # Apply formatting
        worksheet = writer.sheets[sheet_name]
        from openpyxl.styles import numbers, PatternFill, Font
        
        # Orange fill for header row
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        header_font = Font(bold=True)  # Bold text with default black color
        
        # Check if the computed columns already exist, if not, insert them
        header_positions = {worksheet.cell(1, idx).value: idx for idx in range(1, worksheet.max_column + 1)}
        
        # Only insert columns if they don't already exist
        has_actual_cost = "actual_cost" in header_positions
        has_no_solar_cost = "no_solar_cost" in header_positions
        has_net_returns = "net_returns" in header_positions
        
        # Find the position to insert after grid_to_battery_kwh
        grid_to_battery_col = header_positions.get("grid_to_battery_kwh")
        
        if grid_to_battery_col and not has_actual_cost:
            # Insert Actual_Cost column
            actual_cost_col = grid_to_battery_col + 1
            worksheet.insert_cols(actual_cost_col)
            worksheet.cell(row=1, column=actual_cost_col).value = "actual_cost"
            worksheet.cell(row=1, column=actual_cost_col).fill = orange_fill
            worksheet.cell(row=1, column=actual_cost_col).font = header_font
            
            # Add formulas for data rows (import cost - export income)
            for row in range(2, worksheet.max_row + 1):
                formula = f"=C{row}-E{row}"
                worksheet.cell(row=row, column=actual_cost_col).value = formula
                worksheet.cell(row=row, column=actual_cost_col).number_format = '£#,##0.00'
            
            # Update header positions
            header_positions = {worksheet.cell(1, idx).value: idx for idx in range(1, worksheet.max_column + 1)}

        if not has_no_solar_cost:
            # Get current actual_cost position
            actual_cost_col = header_positions.get("actual_cost")
            if actual_cost_col:
                no_solar_cost_col = actual_cost_col + 1
                worksheet.insert_cols(no_solar_cost_col)
                worksheet.cell(row=1, column=no_solar_cost_col).value = "no_solar_cost"
                worksheet.cell(row=1, column=no_solar_cost_col).fill = orange_fill
                worksheet.cell(row=1, column=no_solar_cost_col).font = header_font
                
                # Add formulas
                positions = {worksheet.cell(1, idx).value: idx for idx in range(1, worksheet.max_column + 1)}
                import_col = positions.get("import_kwh")
                pv_col = positions.get("pv_to_home_kwh")
                battery_col = positions.get("grid_to_battery_kwh")
                
                if import_col and pv_col and battery_col:
                    import_col_letter = get_column_letter(import_col)
                    pv_col_letter = get_column_letter(pv_col)
                    battery_col_letter = get_column_letter(battery_col)
                    for row in range(2, worksheet.max_row + 1):
                        formula = f"=(({import_col_letter}{row}+{pv_col_letter}{row})-{battery_col_letter}{row})*{baseline_rate}"
                        worksheet.cell(row=row, column=no_solar_cost_col).value = formula
                        worksheet.cell(row=row, column=no_solar_cost_col).number_format = '£#,##0.00'
                
                # Update header positions
                header_positions = {worksheet.cell(1, idx).value: idx for idx in range(1, worksheet.max_column + 1)}

        if not has_net_returns:
            # Get current no_solar_cost and actual_cost positions
            no_solar_cost_col = header_positions.get("no_solar_cost")
            actual_cost_col = header_positions.get("actual_cost")
            
            if no_solar_cost_col and actual_cost_col:
                roi_col = no_solar_cost_col + 1
                worksheet.insert_cols(roi_col)
                worksheet.cell(row=1, column=roi_col).value = "net_returns"
                worksheet.cell(row=1, column=roi_col).fill = orange_fill
                worksheet.cell(row=1, column=roi_col).font = header_font
                
                no_solar_cost_letter = get_column_letter(no_solar_cost_col)
                actual_cost_letter = get_column_letter(actual_cost_col)
                for row in range(2, worksheet.max_row + 1):
                    formula = f"={no_solar_cost_letter}{row}-{actual_cost_letter}{row}"
                    worksheet.cell(row=row, column=roi_col).value = formula
                    worksheet.cell(row=row, column=roi_col).number_format = '£#,##0.00'
        
        # ===== ENSURE ALL ROWS HAVE FORMULAS (even if columns already exist) =====
        # Update header positions one final time to get all current column positions
        header_positions = {worksheet.cell(1, idx).value: idx for idx in range(1, worksheet.max_column + 1)}
        
        actual_cost_col = header_positions.get("actual_cost")
        no_solar_cost_col = header_positions.get("no_solar_cost")
        net_returns_col = header_positions.get("net_returns")
        import_col = header_positions.get("import_kwh")
        export_income_col = header_positions.get("export_income")
        pv_col = header_positions.get("pv_to_home_kwh")
        battery_col = header_positions.get("grid_to_battery_kwh")
        
        # Set formulas for all data rows
        for row in range(2, worksheet.max_row + 1):
            # Set actual_cost formula: import_cost - export_income
            if actual_cost_col:
                cell = worksheet.cell(row=row, column=actual_cost_col)
                formula = f"=C{row}-E{row}"
                cell.value = formula
                cell.number_format = '£#,##0.00'
            
            # Set no_solar_cost formula: ((import + pv_to_home - battery) * baseline_rate)
            if no_solar_cost_col and import_col and pv_col and battery_col:
                cell = worksheet.cell(row=row, column=no_solar_cost_col)
                import_col_letter = get_column_letter(import_col)
                pv_col_letter = get_column_letter(pv_col)
                battery_col_letter = get_column_letter(battery_col)
                formula = f"=(({import_col_letter}{row}+{pv_col_letter}{row})-{battery_col_letter}{row})*{baseline_rate}"
                cell.value = formula
                cell.number_format = '£#,##0.00'
            
            # Set net_returns formula: no_solar_cost - actual_cost
            if net_returns_col and no_solar_cost_col and actual_cost_col:
                cell = worksheet.cell(row=row, column=net_returns_col)
                no_solar_cost_letter = get_column_letter(no_solar_cost_col)
                actual_cost_letter = get_column_letter(actual_cost_col)
                formula = f"={no_solar_cost_letter}{row}-{actual_cost_letter}{row}"
                cell.value = formula
                cell.number_format = '£#,##0.00'

        # Format header row
        for cell in worksheet[1]:
            cell.fill = orange_fill
            cell.font = header_font
        
        # Format data rows
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                col_header = worksheet.cell(1, cell.column).value
                
                if col_header and "kwh" in col_header.lower():
                    # Format kWh to 1 decimal place
                    cell.number_format = '0.0'
                elif col_header and ("cost" in col_header.lower() or "income" in col_header.lower() 
                                     or "savings" in col_header.lower() or "return" in col_header.lower() 
                                     or col_header == "net_returns"):
                    # Format as British Pounds with 2 decimal places
                    cell.number_format = '£#,##0.00'
        
        # Auto-fit column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set width with some padding
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 for very long content
            worksheet.column_dimensions[column_letter].width = adjusted_width